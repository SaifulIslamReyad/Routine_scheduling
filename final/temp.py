import tkinter as tk
from tkinter import messagebox
import os
import json
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill
from collections import defaultdict
from openpyxl.styles.borders import Border, Side
from tkinter import messagebox, ttk
import os
import json
import tkinter as tk
from tkinter import ttk, messagebox

# File paths
PREF_FILE = "teacher_preferences.json"
RANK_FILE = "teacher_rank.json"
COURSE_FILE = "courses.json"

# Load existing data
preferences = {}
if os.path.exists(PREF_FILE):
    with open(PREF_FILE, "r") as f:
        preferences = json.load(f)

teacher_ranks = {}
if os.path.exists(RANK_FILE):
    with open(RANK_FILE, "r") as f:
        teacher_ranks = json.load(f)

courses = []
if os.path.exists(COURSE_FILE):
    with open(COURSE_FILE, "r") as f:
        courses = json.load(f)

# Constants
DAYS = ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday"]
SLOTS = list(range(1, 8))
slot_timings = {
    1: "9:00-10:00", 2: "10:00-11:00", 3: "11:00-12:00",
    4: "12:00-1:00", 5: "2:00-3:00", 6: "3:00-4:00", 7: "4:00-5:00"
}

# Tkinter setup
root = tk.Tk()
root.title("Teacher Course & Preference Manager")
root.geometry("900x500")
root.configure(bg="#eef5fb")

DEFAULT_FONT = ("Helvetica", 12)
LABEL_FONT = ("Helvetica", 12, "bold")
HEADER_FONT = ("Helvetica", 16, "bold")

title = tk.Label(root, text="Teacher Course & Preference Entry", font=HEADER_FONT,
                 bg="#eef5fb", fg="#003366")
title.pack(pady=10)

notebook = ttk.Notebook(root)
notebook.pack(fill="both", expand=True, padx=10, pady=10)

# --------------------------------------
# 1. Teacher Info Tab
# --------------------------------------

def save_teacher():
    name = teacher_name_entry.get().strip()
    rank = teacher_rank_entry.get().strip()
    if not name or not rank:
        messagebox.showerror("Error", "Please enter both teacher name and rank.")
        return
    try:
        rank = int(rank)
    except ValueError:
        messagebox.showerror("Error", "Rank must be an integer.")
        return
    teacher_ranks[name] = rank
    with open(RANK_FILE, "w") as f:
        json.dump(teacher_ranks, f, indent=2)
    teacher_name_entry.delete(0, tk.END)
    teacher_rank_entry.delete(0, tk.END)
    update_dropdowns()
    show_teacher_list()
    messagebox.showinfo("Saved", f"Saved teacher: {name}")

def show_teacher_list():
    teacher_listbox.delete(0, tk.END)
    for name, rank in sorted(teacher_ranks.items(), key=lambda x: x[1]):
        teacher_listbox.insert(tk.END, f"{name} (Rank: {rank})")

teacher_tab = tk.Frame(notebook, bg="#f4f9ff")
notebook.add(teacher_tab, text="1. Teacher Info")

tk.Label(teacher_tab, text="Teacher Name:", font=LABEL_FONT, bg="#f4f9ff").grid(row=0, column=0, padx=10, pady=10, sticky="e")
teacher_name_entry = tk.Entry(teacher_tab, width=30, font=DEFAULT_FONT)
teacher_name_entry.grid(row=0, column=1, padx=5)

tk.Label(teacher_tab, text="Rank:", font=LABEL_FONT, bg="#f4f9ff").grid(row=0, column=2, padx=10, pady=10, sticky="e")
teacher_rank_entry = tk.Entry(teacher_tab, width=10, font=DEFAULT_FONT)
teacher_rank_entry.grid(row=0, column=3, padx=5)

tk.Button(teacher_tab, text="Save Teacher", command=save_teacher, bg="#0066cc", fg="white",
          padx=15, pady=6, font=DEFAULT_FONT).grid(row=1, columnspan=4, pady=15)

teacher_listbox = tk.Listbox(teacher_tab, font=DEFAULT_FONT, width=60)
teacher_listbox.grid(row=2, columnspan=4, pady=10)
show_teacher_list()

# --------------------------------------
# 2. Course Entry Tab
# --------------------------------------

course_teacher_var = tk.StringVar()

def save_course():
    teacher = course_teacher_var.get()
    title = course_title_entry.get().strip()
    code = course_code_entry.get().strip()
    credit = course_credit_entry.get().strip()
    year = course_year_entry.get().strip()
    if not teacher or not title or not code or not credit or not year:
        messagebox.showerror("Error", "Please fill in all fields.")
        return
    try:
        credit = float(credit)
        year = int(year)
    except ValueError:
        messagebox.showerror("Error", "Invalid year or credit.")
        return
    courses.append({
        "year": year,
        "code": code,
        "title": title,
        "credit": credit,
        "teacher": teacher
    })
    with open(COURSE_FILE, "w") as f:
        json.dump(courses, f, indent=2)
    course_title_entry.delete(0, tk.END)
    course_code_entry.delete(0, tk.END)
    course_credit_entry.delete(0, tk.END)
    course_year_entry.delete(0, tk.END)
    update_dropdowns()
    show_course_list()
    messagebox.showinfo("Saved", f"Course saved for {teacher}")

def show_course_list():
    course_listbox.delete(0, tk.END)
    for c in courses:
        course_listbox.insert(tk.END, f"{c['code']} -  ({c['credit']} cr) [{c['teacher']}]")

course_tab = tk.Frame(notebook, bg="#f4f9ff")
notebook.add(course_tab, text="2. Course Entry")

tk.Label(course_tab, text="Teacher Name:", font=LABEL_FONT, bg="#f4f9ff").grid(row=0, column=0, padx=10, pady=10)
course_teacher_dropdown = ttk.Combobox(course_tab, textvariable=course_teacher_var, state="readonly", width=27, font=DEFAULT_FONT)
course_teacher_dropdown.grid(row=0, column=1, padx=5)

tk.Label(course_tab, text="Course Title:", font=LABEL_FONT, bg="#f4f9ff").grid(row=1, column=0, padx=10, pady=10)
course_title_entry = tk.Entry(course_tab, width=20, font=DEFAULT_FONT)
course_title_entry.grid(row=1, column=1)

tk.Label(course_tab, text="Course Code:", font=LABEL_FONT, bg="#f4f9ff").grid(row=1, column=2, padx=10)
course_code_entry = tk.Entry(course_tab, width=20, font=DEFAULT_FONT)
course_code_entry.grid(row=1, column=3)

tk.Label(course_tab, text="Credit:", font=LABEL_FONT, bg="#f4f9ff").grid(row=2, column=0, padx=10, pady=10)
course_credit_entry = tk.Entry(course_tab, width=20, font=DEFAULT_FONT)
course_credit_entry.grid(row=2, column=1)

tk.Label(course_tab, text="Year:", font=LABEL_FONT, bg="#f4f9ff").grid(row=2, column=2, padx=10)
course_year_entry = tk.Entry(course_tab, width=20, font=DEFAULT_FONT)
course_year_entry.grid(row=2, column=3)

tk.Button(course_tab, text="Save Course", command=save_course, bg="#0066cc", fg="white",
          padx=15, pady=6, font=DEFAULT_FONT).grid(row=3, columnspan=4, pady=15)

course_listbox = tk.Listbox(course_tab, font=DEFAULT_FONT, width=80)
course_listbox.grid(row=4, columnspan=4, pady=10)
show_course_list()

# --------------------------------------
# 3. Preferences Tab
# --------------------------------------

pref_teacher_var = tk.StringVar()
slot_buttons = {}
selected_slots = set()

def toggle_slot(day, slot):
    key = (day, slot)
    if key in selected_slots:
        selected_slots.remove(key)
        slot_buttons[key].config(bg="SystemButtonFace")
    else:
        selected_slots.add(key)
        slot_buttons[key].config(bg="lightgreen")

def save_preferences():
    name = pref_teacher_var.get()
    if not name:
        messagebox.showerror("Error", "Select a teacher.")
        return
    prefs = sorted(list(selected_slots), key=lambda x: (DAYS.index(x[0]), x[1]))
    preferences[name] = prefs
    with open(PREF_FILE, "w") as f:
        json.dump(preferences, f, indent=2)
    messagebox.showinfo("Saved", f"Preferences saved for {name}")
    for key in selected_slots.copy():
        slot_buttons[key].config(bg="SystemButtonFace")
    selected_slots.clear()

pref_tab = tk.Frame(notebook, bg="#f4f9ff")
notebook.add(pref_tab, text="3. Preferences")

tk.Label(pref_tab, text="Select Teacher:", bg="#f4f9ff").grid(row=0, column=0, padx=10, pady=10)
pref_teacher_dropdown = ttk.Combobox(pref_tab, textvariable=pref_teacher_var, state="readonly", width=25)
pref_teacher_dropdown.grid(row=0, column=1, padx=5)

slot_frame = tk.LabelFrame(pref_tab, text="Preferred Slots", bg="#e8f0fe", font=("Helvetica", 10))
slot_frame.grid(row=1, column=0, columnspan=4, padx=10, pady=10)

tk.Label(slot_frame, text="", bg="#e8f0fe").grid(row=0, column=0)
for j, slot in enumerate(SLOTS):
    tk.Label(slot_frame, text=slot_timings[slot], bg="#e8f0fe").grid(row=0, column=j+1)

for i, day in enumerate(DAYS):
    tk.Label(slot_frame, text=day, bg="#e8f0fe").grid(row=i+1, column=0)
    for j, slot in enumerate(SLOTS):
        btn = tk.Button(slot_frame, text=str(slot), width=6)
        btn.grid(row=i+1, column=j+1, padx=2, pady=2)
        btn.config(command=lambda d=day, s=slot: toggle_slot(d, s))
        slot_buttons[(day, slot)] = btn

tk.Button(pref_tab, text="Save Preferences", command=save_preferences, bg="#0066cc", fg="white",
          padx=20, pady=5).grid(row=2, columnspan=4, pady=10)

# --------------------------------------
# Update dropdowns and run
# --------------------------------------

def update_dropdowns():
    teacher_list = list(teacher_ranks.keys())
    course_teacher_dropdown["values"] = teacher_list
    pref_teacher_dropdown["values"] = teacher_list

update_dropdowns()
root.mainloop()


# ///////////////////////////

with open("courses.json", "r") as f:courses = json.load(f)
with open("teacher_rank.json", "r") as f:teacher_rank = json.load(f)
with open("teacher_preferences.json", "r") as f:teacher_preferences = json.load(f)
days = ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday"]
routine = {year: {day: {slot: None for slot in range(1, 8)} for day in days}for year in [1, 2, 3, 4]}

teacher_schedule = defaultdict(lambda: {day: set() for day in days})
courses.sort(key=lambda c: teacher_rank.get(c["teacher"], float("inf")))


def can_assign(teacher, year, day, slot):  return ( routine[year][day][slot] is None and slot not in teacher_schedule[teacher][day] )
def reassign_slot(year, day, slot, requesting_teacher):
    assignment = routine[year][day][slot]
    if assignment is None or assignment[1] == requesting_teacher:
        return True  

    course_code, current_teacher = assignment
    prefs = teacher_preferences.get(current_teacher, [])

    for new_day, new_slot in prefs:
        if new_day == day and new_slot == slot:
            continue

        if can_assign(current_teacher, year, new_day, new_slot):
            _move_assignment(year, day, slot, new_day, new_slot, course_code, current_teacher)
            return True

        if reassign_slot(year, new_day, new_slot, current_teacher):
            if can_assign(current_teacher, year, new_day, new_slot):
                _move_assignment(year, day, slot, new_day, new_slot, course_code, current_teacher)
                return True

    return False 

def _move_assignment(year, old_day, old_slot, new_day, new_slot, course_code, teacher):
    if int(course_code[-1])%2==1:
        routine[year][old_day][old_slot] = None
        teacher_schedule[teacher][old_day].remove(old_slot)
        routine[year][new_day][new_slot] = (course_code, teacher)
        teacher_schedule[teacher][new_day].add(new_slot)
        
    else:
        routine[year][old_day][old_slot] = None
        routine[year][old_day][old_slot+1] = None
        routine[year][old_day][old_slot+2] = None

        teacher_schedule[teacher][old_day].remove(old_slot)
        teacher_schedule[teacher][old_day].remove(old_slot+1)
        teacher_schedule[teacher][old_day].remove(old_slot+2)

        routine[year][new_day][new_slot] = (course_code, teacher)
        routine[year][new_day][new_slot+1] = (course_code, teacher)
        routine[year][new_day][new_slot+2] = (course_code, teacher)

        teacher_schedule[teacher][new_day].add(new_slot)
        teacher_schedule[teacher][new_day].add(new_slot+1)
        teacher_schedule[teacher][new_day].add(new_slot+2)



def assign_course(course):
    year = course["year"]
    code = course["code"]
    teacher = course["teacher"]
    credits = course["credit"]
    assigned = 0
    prefs = teacher_preferences.get(teacher, [])

    for day, slot in prefs:
        if assigned >= credits:
            return

        if can_assign(teacher, year, day, slot):
            if int(code[-1])%2==1:
                routine[year][day][slot] = (code, teacher)
                teacher_schedule[teacher][day].add(slot)
                assigned += 1
            else:
                routine[year][day][slot] = (code, teacher)
                routine[year][day][slot+1] = (code, teacher)
                routine[year][day][slot+2] = (code, teacher)
                teacher_schedule[teacher][day].add(slot)
                teacher_schedule[teacher][day].add(slot+1)
                teacher_schedule[teacher][day].add(slot+2)
                assigned += 1                


    for day, slot in prefs:
        if assigned >= credits:
            return
        
        if reassign_slot(year,day, slot, teacher):
            if int(code[-1])%2==1:
                routine[year][day][slot] = (code, teacher)
                teacher_schedule[teacher][day].add(slot)
                assigned += 1
            else:
                routine[year][day][slot] = (code, teacher)
                routine[year][day][slot+1] = (code, teacher)
                routine[year][day][slot+2] = (code, teacher)
                teacher_schedule[teacher][day].add(slot)
                teacher_schedule[teacher][day].add(slot+1)
                teacher_schedule[teacher][day].add(slot+2)
                assigned += 1 

    if assigned < credits:
        print(f"⚠ Warning: Could not fully assign {code} for {teacher}!")



# ////////////////////////////////////////////////////////////////////////////////////////////////
import openpyxl
from openpyxl.styles import Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

def printing():
    slot_timings = {
        1: "9:00-10:00", 2: "10:00-11:00", 3: "11:00-12:00",
        4: "12:00-1:00", "Break": "1:00-2:00", 5: "2:00-3:00",
        6: "3:00-4:00", 7: "4:00-5:00"
    }
    year_colors = {1: "D9E1F2", 2: "E2EFDA", 3: "FFF2CC", 4: "FCE4D6"}
    free_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Routine"

    ws.cell(row=1, column=1, value="Day").alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=1, column=2, value="Year").alignment = Alignment(horizontal="center", vertical="center")

    col_index = 3
    for slot in [1, 2, 3, 4, "Break", 5, 6, 7]:
        cell = ws.cell(row=1, column=col_index, value=slot_timings[slot])
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        col_index += 1

    current_row = 2
    for day in days:
        start_row = current_row
        for year in [1, 2, 3, 4]:
            year_name = (
                f"{year}st Year" if year == 1 else
                f"{year}nd Year" if year == 2 else
                f"{year}rd Year" if year == 3 else
                f"{year}th Year"
            )
            year_cell = ws.cell(row=current_row, column=2, value=year_name)
            year_cell.fill = PatternFill(start_color=year_colors[year], end_color=year_colors[year], fill_type="solid")
            year_cell.alignment = Alignment(horizontal="center", vertical="center")
            year_cell.border = thin_border

            col_index = 3
            for slot in [1, 2, 3, 4, "Break", 5, 6, 7]:
                if slot == "Break":
                    value = "BREAK"
                    cell = ws.cell(row=current_row, column=col_index, value=value)
                    cell.fill = free_fill
                else:
                    entry = routine[year][day][slot]
                    if entry:
                        code, teacher = entry
                        value = f"{code} ({teacher})"
                    else:
                        value = "-"
                    cell = ws.cell(row=current_row, column=col_index, value=value)
                    if value.strip() == "-":
                        cell.fill = free_fill
                    else:
                        cell.fill = PatternFill(start_color=year_colors[year], end_color=year_colors[year], fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border
                col_index += 1

            current_row += 1

        ws.merge_cells(start_row=start_row, start_column=1, end_row=current_row - 1, end_column=1)
        day_cell = ws.cell(row=start_row, column=1, value=day)
        day_cell.alignment = Alignment(horizontal="center", vertical="center")
        day_cell.border = thin_border

        current_row += 1

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        col = 3
        while col <= 10:
            start_col = col
            value = row[col - 1].value
            while col + 1 <= 10 and row[col].value == value and value not in ["-", "BREAK"]:
                col += 1
            if col > start_col:
                ws.merge_cells(
                    start_row=row[0].row, start_column=start_col,
                    end_row=row[0].row, end_column=col
                )
                merged_cell = ws.cell(row=row[0].row, column=start_col)
                merged_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            col += 1

    for col in range(1, 11):
        ws.column_dimensions[get_column_letter(col)].width = 22

    wb.save("routine_final.xlsx")
    print("\n 💕 Final Routine with Time Slots (including Break) saved as 'routine_final.xlsx'!")

for course in courses:
    assign_course(course)


printing()


