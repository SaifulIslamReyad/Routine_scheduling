import tkinter as tk
from tkinter import messagebox
import os
import json

# File paths
PREF_FILE = "teacher_preferences.json"
RANK_FILE = "teacher_rank.json"
COURSE_FILE = "courses.json"

# Load existing data
preferences = {}
if os.path.exists(PREF_FILE):
    with open(PREF_FILE, "r") as f:
        preferences = json.load(f)

teacher_ranks = {}
if os.path.exists(RANK_FILE):
    with open(RANK_FILE, "r") as f:
        teacher_ranks = json.load(f)

courses = []
if os.path.exists(COURSE_FILE):
    with open(COURSE_FILE, "r") as f:
        courses = json.load(f)

# Constants
DAYS = ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday"]
SLOTS = list(range(1, 8))
slot_timings = {
    1: "9:00-10:00", 2: "10:00-11:00", 3: "11:00-12:00",
    4: "12:00-1:00", 5: "2:00-3:00", 6: "3:00-4:00", 7: "4:00-5:00"
}

# Main UI
root = tk.Tk()
root.title("Teacher Preferences and Course Input")

# Inputs: Teacher name & rank
tk.Label(root, text="Teacher Name:").grid(row=0, column=0)
name_entry = tk.Entry(root)
name_entry.grid(row=0, column=1)

tk.Label(root, text="Rank:").grid(row=0, column=2)
rank_entry = tk.Entry(root, width=5)
rank_entry.grid(row=0, column=3)

# Inputs: Course info
tk.Label(root, text="Year:").grid(row=1, column=0)
year_entry = tk.Entry(root, width=5)
year_entry.grid(row=1, column=1)

tk.Label(root, text="Course:").grid(row=1, column=2)
course_entry = tk.Entry(root)
course_entry.grid(row=1, column=3)

tk.Label(root, text="Code:").grid(row=1, column=4)
code_entry = tk.Entry(root)
code_entry.grid(row=1, column=5)

tk.Label(root, text="Credit:").grid(row=1, column=6)
credit_entry = tk.Entry(root, width=5)
credit_entry.grid(row=1, column=7)

# Slot header row
tk.Label(root, text="").grid(row=2, column=0)
for j, slot in enumerate(SLOTS):
    tk.Label(root, text=slot_timings[slot], font=("Arial", 9, "bold")).grid(row=2, column=j + 1)

selected_cells = set()
buttons = {}

# Time slot buttons
def toggle_cell(day, slot, btn):
    key = (day, slot)
    if key in selected_cells:
        selected_cells.remove(key)
        btn.config(bg="SystemButtonFace")
    else:
        selected_cells.add(key)
        btn.config(bg="lightgreen")

for i, day in enumerate(DAYS):
    tk.Label(root, text=day).grid(row=i + 3, column=0)
    for j, slot in enumerate(SLOTS):
        btn = tk.Button(root, text=str(slot), width=6,
                        command=lambda d=day, s=slot: toggle_cell(d, s, buttons[(d, s)]))
        btn.grid(row=i + 3, column=j + 1)
        buttons[(day, slot)] = btn

# Clear fields
def clear_all():
    name_entry.delete(0, tk.END)
    rank_entry.delete(0, tk.END)
    year_entry.delete(0, tk.END)
    course_entry.delete(0, tk.END)
    code_entry.delete(0, tk.END)
    credit_entry.delete(0, tk.END)
    for key in selected_cells.copy():
        buttons[key].config(bg="SystemButtonFace")
    selected_cells.clear()

# Save all data
def save_all():
    name = name_entry.get().strip()
    rank = rank_entry.get().strip()
    year = year_entry.get().strip()
    course = course_entry.get().strip()
    code = code_entry.get().strip()
    credit = credit_entry.get().strip()

    if not name or not rank:
        messagebox.showerror("Error", "Enter both teacher name and rank.")
        return

    try:
        rank = int(rank)
    except ValueError:
        messagebox.showerror("Error", "Rank must be an integer.")
        return

    # Save teacher rank
    teacher_ranks[name] = rank
    with open(RANK_FILE, "w") as f:
        json.dump(teacher_ranks, f, indent=2)

    # Save preferences
    if selected_cells:
        sorted_prefs = sorted(list(selected_cells), key=lambda x: (DAYS.index(x[0]), x[1]))
        preferences[name] = sorted_prefs
        with open(PREF_FILE, "w") as f:
            json.dump(preferences, f, indent=2)

    # Save course info
    if year and course and code and credit:
        try:
            year = int(year)
            credit = float(credit)
            courses.append({
                "year": year,
                "code": code,
                "credit": credit,
                "teacher": name
            })
            with open(COURSE_FILE, "w") as f:
                json.dump(courses, f, indent=2)
        except ValueError:
            messagebox.showerror("Error", "Invalid year or credit format.")
            return

    messagebox.showinfo("Success", f"All data saved for {name}")
    clear_all()

# Save button
tk.Button(root, text="Save All", command=save_all).grid(row=10, column=0, columnspan=8, pady=10)

root.mainloop()

# ///////////////////////////

with open("courses.json", "r") as f:courses = json.load(f)
with open("teacher_rank.json", "r") as f:teacher_rank = json.load(f)
with open("teacher_preferences.json", "r") as f:teacher_preferences = json.load(f)
days = ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday"]
routine = {year: {day: {slot: None for slot in range(1, 8)} for day in days}for year in [1, 2, 3, 4]}

teacher_schedule = defaultdict(lambda: {day: set() for day in days})
courses.sort(key=lambda c: teacher_rank.get(c["teacher"], float("inf")))


def can_assign(teacher, year, day, slot):  return ( routine[year][day][slot] is None and slot not in teacher_schedule[teacher][day] )
def reassign_slot(year, day, slot, requesting_teacher):
    assignment = routine[year][day][slot]
    if assignment is None or assignment[1] == requesting_teacher:
        return True  

    course_code, current_teacher = assignment
    prefs = teacher_preferences.get(current_teacher, [])

    for new_day, new_slot in prefs:
        if new_day == day and new_slot == slot:
            continue

        if can_assign(current_teacher, year, new_day, new_slot):
            _move_assignment(year, day, slot, new_day, new_slot, course_code, current_teacher)
            return True

        if reassign_slot(year, new_day, new_slot, current_teacher):
            if can_assign(current_teacher, year, new_day, new_slot):
                _move_assignment(year, day, slot, new_day, new_slot, course_code, current_teacher)
                return True

    return False 

def _move_assignment(year, old_day, old_slot, new_day, new_slot, course_code, teacher):
    if int(course_code[-1])%2==1:
        routine[year][old_day][old_slot] = None
        teacher_schedule[teacher][old_day].remove(old_slot)
        routine[year][new_day][new_slot] = (course_code, teacher)
        teacher_schedule[teacher][new_day].add(new_slot)
        
    else:
        routine[year][old_day][old_slot] = None
        routine[year][old_day][old_slot+1] = None
        routine[year][old_day][old_slot+2] = None

        teacher_schedule[teacher][old_day].remove(old_slot)
        teacher_schedule[teacher][old_day].remove(old_slot+1)
        teacher_schedule[teacher][old_day].remove(old_slot+2)

        routine[year][new_day][new_slot] = (course_code, teacher)
        routine[year][new_day][new_slot+1] = (course_code, teacher)
        routine[year][new_day][new_slot+2] = (course_code, teacher)

        teacher_schedule[teacher][new_day].add(new_slot)
        teacher_schedule[teacher][new_day].add(new_slot+1)
        teacher_schedule[teacher][new_day].add(new_slot+2)



def assign_course(course):
    year = course["year"]
    code = course["code"]
    teacher = course["teacher"]
    credits = course["credit"]
    assigned = 0
    prefs = teacher_preferences.get(teacher, [])

    for day, slot in prefs:
        if assigned >= credits:
            return

        if can_assign(teacher, year, day, slot):
            if int(code[-1])%2==1:
                routine[year][day][slot] = (code, teacher)
                teacher_schedule[teacher][day].add(slot)
                assigned += 1
            else:
                routine[year][day][slot] = (code, teacher)
                routine[year][day][slot+1] = (code, teacher)
                routine[year][day][slot+2] = (code, teacher)
                teacher_schedule[teacher][day].add(slot)
                teacher_schedule[teacher][day].add(slot+1)
                teacher_schedule[teacher][day].add(slot+2)
                assigned += 1                


    for day, slot in prefs:
        if assigned >= credits:
            return
        
        if reassign_slot(year,day, slot, teacher):
            if int(code[-1])%2==1:
                routine[year][day][slot] = (code, teacher)
                teacher_schedule[teacher][day].add(slot)
                assigned += 1
            else:
                routine[year][day][slot] = (code, teacher)
                routine[year][day][slot+1] = (code, teacher)
                routine[year][day][slot+2] = (code, teacher)
                teacher_schedule[teacher][day].add(slot)
                teacher_schedule[teacher][day].add(slot+1)
                teacher_schedule[teacher][day].add(slot+2)
                assigned += 1 

    if assigned < credits:
        print(f"⚠ Warning: Could not fully assign {code} for {teacher}!")

def lab_assign_course(course):
    pass 

# ////////////////////////////////////////////////////////////////////////////////////////////////
import openpyxl
from openpyxl.styles import Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

def printing():
    slot_timings = {
        1: "9:00-10:00", 2: "10:00-11:00", 3: "11:00-12:00",
        4: "12:00-1:00", "Break": "1:00-2:00", 5: "2:00-3:00",
        6: "3:00-4:00", 7: "4:00-5:00"
    }
    year_colors = {1: "D9E1F2", 2: "E2EFDA", 3: "FFF2CC", 4: "FCE4D6"}
    free_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Routine"

    ws.cell(row=1, column=1, value="Day").alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=1, column=2, value="Year").alignment = Alignment(horizontal="center", vertical="center")

    col_index = 3
    for slot in [1, 2, 3, 4, "Break", 5, 6, 7]:
        cell = ws.cell(row=1, column=col_index, value=slot_timings[slot])
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        col_index += 1

    current_row = 2
    for day in days:
        start_row = current_row
        for year in [1, 2, 3, 4]:
            year_name = (
                f"{year}st Year" if year == 1 else
                f"{year}nd Year" if year == 2 else
                f"{year}rd Year" if year == 3 else
                f"{year}th Year"
            )
            year_cell = ws.cell(row=current_row, column=2, value=year_name)
            year_cell.fill = PatternFill(start_color=year_colors[year], end_color=year_colors[year], fill_type="solid")
            year_cell.alignment = Alignment(horizontal="center", vertical="center")
            year_cell.border = thin_border

            col_index = 3
            for slot in [1, 2, 3, 4, "Break", 5, 6, 7]:
                if slot == "Break":
                    value = "BREAK"
                    cell = ws.cell(row=current_row, column=col_index, value=value)
                    cell.fill = free_fill
                else:
                    entry = routine[year][day][slot]
                    if entry:
                        code, teacher = entry
                        value = f"{code} ({teacher})"
                    else:
                        value = "-"
                    cell = ws.cell(row=current_row, column=col_index, value=value)
                    if value.strip() == "-":
                        cell.fill = free_fill
                    else:
                        cell.fill = PatternFill(start_color=year_colors[year], end_color=year_colors[year], fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border
                col_index += 1

            current_row += 1

        ws.merge_cells(start_row=start_row, start_column=1, end_row=current_row - 1, end_column=1)
        day_cell = ws.cell(row=start_row, column=1, value=day)
        day_cell.alignment = Alignment(horizontal="center", vertical="center")
        day_cell.border = thin_border

        current_row += 1

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        col = 3
        while col <= 10:
            start_col = col
            value = row[col - 1].value
            while col + 1 <= 10 and row[col].value == value and value not in ["-", "BREAK"]:
                col += 1
            if col > start_col:
                ws.merge_cells(
                    start_row=row[0].row, start_column=start_col,
                    end_row=row[0].row, end_column=col
                )
                merged_cell = ws.cell(row=row[0].row, column=start_col)
                merged_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            col += 1

    for col in range(1, 11):
        ws.column_dimensions[get_column_letter(col)].width = 22

    wb.save("routine_final.xlsx")
    print("\n 💕 Final Routine with Time Slots (including Break) saved as 'routine_final.xlsx'!")

for course in courses:
    assign_course(course)


printing()